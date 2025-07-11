import json
import pandas as pd
import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 输入路径
input_path = "results/error_GEN.json"
output_path = os.path.splitext(input_path)[0] + "_sorted_merged.xlsx"

# 加载 JSON 文件
with open(input_path, "r") as f:
    raw_data = json.load(f)

# === 计算 average_recall 并整理数据 ===
flattened = []

for error_type, subtypes in raw_data.items():
    sub_items = [(k, v) for k, v in subtypes.items() if k != "average_recall"]
    total_recall = 0
    for _, v in sub_items:
        total_recall += v["recall"]
    avg_recall = round(total_recall / len(sub_items), 4) if sub_items else 0
    raw_data[error_type]["average_recall"] = avg_recall

# === 排序大类 ===
sorted_data = dict(sorted(raw_data.items(), key=lambda item: item[1]["average_recall"], reverse=True))

# === 生成 DataFrame 数据 ===
records = []
for error_type, subtypes in sorted_data.items():
    avg_recall = subtypes["average_recall"]
    sub_items = [(k, v) for k, v in subtypes.items() if k != "average_recall"]
    for i, (sub_error_type, metrics) in enumerate(sub_items):
        is_last = i == len(sub_items) - 1
        records.append({
            "Error Type": error_type,
            "Sub Error Type": sub_error_type,
            "Total": metrics.get("total", 0),
            "Correctly Detected Errors": metrics.get("detected_false_positives", 0),
            "Recall": metrics.get("recall", 0),
            "Average Recall (Error Type)": avg_recall if is_last else None
        })

# 导出到 Excel
df = pd.DataFrame(records)
df.to_excel(output_path, index=False)

# === 合并单元格 ===
wb = load_workbook(output_path)
ws = wb.active

current_type = None
start_row = 2
for i, row in enumerate(df["Error Type"], start=2):
    if row != current_type:
        if current_type is not None:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=i-1, end_column=1)
        current_type = row
        start_row = i
# 最后一组也要合并
ws.merge_cells(start_row=start_row, start_column=1, end_row=len(df)+1, end_column=1)

wb.save(output_path)
print(f"✅ 成功导出（已排序和合并）：{output_path}")
